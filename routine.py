"""
Automated routine — Task 2
1. Reads template files to observe their column structure.
2. Reads every sheet in Input.xlsx (read fidelity checkpoint).
3. For each sheet, normalizes phones, silently deduplicates, then builds
   campaign and customers output files from the same valid row set.
   Bad-phone rows go to .review files.

Idempotency design:
    This routine runs on a schedule (poll), not on a file-change event,
    because SharePoint connectors do not expose a "file created/updated"
    trigger. To avoid reprocessing unchanged input on every scheduled run,
    a SHA-256 hash of Input.xlsx is stored in processed/Input.xlsx.hash
    after each successful run. On the next run, if the hash matches the
    stored value the routine exits early. If the file has changed (or no
    marker exists yet), the routine processes normally and updates the marker.

Prerequisites:
    pip install openpyxl pandas

Usage:
    python routine.py
"""

import hashlib
import re
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl import Workbook

REPO               = Path(__file__).parent
INPUT_FILE         = REPO / "Input.xlsx"
CAMPAIGN_TEMPLATE  = REPO / "campaign_Otros_Proyectos.xlsx"
CUSTOMERS_TEMPLATE = REPO / "customers_Otros_Proyectos.xlsx"
OUTPUT_DIR         = REPO / "output"
PROCESSED_DIR      = REPO / "processed"
HASH_FILE          = PROCESSED_DIR / "Input.xlsx.hash"


# ---------------------------------------------------------------------------
# Idempotency helpers
# ---------------------------------------------------------------------------

def file_hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def already_processed(path: Path) -> bool:
    if not HASH_FILE.exists():
        return False
    return HASH_FILE.read_text().strip() == file_hash(path)


def mark_processed(path: Path) -> None:
    PROCESSED_DIR.mkdir(exist_ok=True)
    HASH_FILE.write_text(file_hash(path))


# ---------------------------------------------------------------------------
# CHECKPOINT 0 — Read template structures
# ---------------------------------------------------------------------------

def read_template_headers(path: Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.worksheets[0]
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    wb.close()
    print(f"  Template '{path.name}': {headers}")
    return headers


def load_voice_model_lookup(path: Path) -> dict[str, str]:
    """Build normalized_phone → voice_model_selection lookup from customers template."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    lookup = {}
    for row in rows[1:]:
        phone_raw, voice = row[0], row[4]
        if phone_raw is None or voice is None:
            continue
        phone, valid = normalize_phone(phone_raw)
        if valid:
            lookup[phone] = voice
    return lookup


# ---------------------------------------------------------------------------
# CHECKPOINT 1 — Read fidelity on Input.xlsx
# ---------------------------------------------------------------------------

def validate_fidelity(path: Path) -> dict[str, pd.DataFrame]:
    if not path.exists():
        raise FileNotFoundError(
            f"Input file not found: {path}\n"
            "Download Input.xlsx from SharePoint and place it next to routine.py"
        )

    wb = openpyxl.load_workbook(path, read_only=True)
    sheets: dict[str, pd.DataFrame] = {}

    print(f"\n=== Checkpoint 1: Read fidelity ===")
    print(f"File: {path.name}  ({path.stat().st_size:,} bytes)")
    print(f"Worksheets found ({len(wb.sheetnames)}): {wb.sheetnames}\n")

    for name in wb.sheetnames:
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        assert rows, f"Sheet '{name}' is empty"

        headers = list(rows[0])
        assert all(h is not None for h in headers), (
            f"Blank header cell in sheet '{name}': {headers}"
        )

        df = pd.DataFrame(rows[1:], columns=headers).dropna(how="all")
        sheets[name] = df

        print(f"  ✅ '{name}'")
        print(f"     Headers ({len(headers)}): {headers}")
        print(f"     Data rows: {len(df)}\n")

    wb.close()
    return sheets


# ---------------------------------------------------------------------------
# Phone normalization
# ---------------------------------------------------------------------------

def normalize_phone(raw) -> tuple[str | None, bool]:
    """
    Returns (normalized_digits, valid).
    - Convert float to int first to avoid '573124457430.0' adding a spurious digit.
    - Strip all non-digit characters.
    - 0057 prefix → strip the leading 00.
    - 12 digits starting with 573 → valid, return as-is.
    - 10 digits starting with 3   → prepend 57, return.
    - Everything else             → invalid, route to .review.
    """
    if pd.isna(raw) or str(raw).strip() in ("", "None"):
        return None, False

    if isinstance(raw, float) and raw.is_integer():
        raw = int(raw)

    digits = re.sub(r"\D", "", str(raw))

    if digits.startswith("0057"):
        digits = digits[2:]

    if len(digits) == 12 and digits.startswith("573"):
        return digits, True

    if len(digits) == 10 and digits.startswith("3"):
        return "57" + digits, True

    return None, False


# ---------------------------------------------------------------------------
# CHECKPOINT 2 — Pre-process, map columns, build output files
# ---------------------------------------------------------------------------

def slugify(name: str) -> str:
    return name.replace(" ", "_")


def preprocess(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Normalize phones and deduplicate (silently drop duplicates).
    Returns (valid_df, bad_df).
    Both campaign and customers are built from the same valid_df,
    guaranteeing equal row counts.
    """
    valid_rows, bad_rows = [], []
    seen: set[str] = set()

    for _, row in df.iterrows():
        phone, valid = normalize_phone(row["Números"])

        if not valid:
            bad_rows.append(row)
            continue

        if phone in seen:
            continue  # silently drop duplicate

        seen.add(phone)
        row = row.copy()
        row["_phone"] = phone
        valid_rows.append(row)

    valid_df = pd.DataFrame(valid_rows).reset_index(drop=True) if valid_rows else pd.DataFrame(columns=list(df.columns) + ["_phone"])
    bad_df   = pd.DataFrame(bad_rows).reset_index(drop=True)   if bad_rows  else pd.DataFrame(columns=list(df.columns))
    return valid_df, bad_df


def build_campaign(valid_df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for _, row in valid_df.iterrows():
        nombre    = str(row["Nombre"])    if pd.notna(row["Nombre"])    else ""
        apellidos = str(row["Apellidos"]) if pd.notna(row["Apellidos"]) else ""
        records.append({
            "number":          row["_phone"],
            "nombre_cliente":  f"{nombre} {apellidos}".strip(),
            "hubspot_deal_id": str(int(row["Negocio ID"])) if pd.notna(row["Negocio ID"]) else "",
        })
    return pd.DataFrame(records)


def build_customers(valid_df: pd.DataFrame, voice_lookup: dict[str, str]) -> pd.DataFrame:
    records = []
    for _, row in valid_df.iterrows():
        records.append({
            "phone":                 row["_phone"],
            "firstname":             row["Nombre"]    if pd.notna(row["Nombre"])    else "",
            "lastname":              row["Apellidos"] if pd.notna(row["Apellidos"]) else "",
            "email":                 None,
            "voice_model_selection": voice_lookup.get(row["_phone"], "Antioquia"),
        })
    return pd.DataFrame(records)


def build_campaign_review(bad_df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for _, row in bad_df.iterrows():
        nombre    = str(row["Nombre"])    if pd.notna(row["Nombre"])    else ""
        apellidos = str(row["Apellidos"]) if pd.notna(row["Apellidos"]) else ""
        raw_phone = str(int(row["Números"])) if isinstance(row["Números"], float) and row["Números"].is_integer() else (str(row["Números"]) if pd.notna(row["Números"]) else "")
        records.append({
            "number":          raw_phone,
            "nombre_cliente":  f"{nombre} {apellidos}".strip(),
            "hubspot_deal_id": str(int(row["Negocio ID"])) if pd.notna(row["Negocio ID"]) else "",
        })
    return pd.DataFrame(records)


def build_customers_review(bad_df: pd.DataFrame) -> pd.DataFrame:
    records = []
    for _, row in bad_df.iterrows():
        raw_phone = str(int(row["Números"])) if isinstance(row["Números"], float) and row["Números"].is_integer() else (str(row["Números"]) if pd.notna(row["Números"]) else "")
        records.append({
            "phone":                 raw_phone,
            "firstname":             row["Nombre"]    if pd.notna(row["Nombre"])    else "",
            "lastname":              row["Apellidos"] if pd.notna(row["Apellidos"]) else "",
            "email":                 None,
            "voice_model_selection": None,
        })
    return pd.DataFrame(records)


def save_excel(df: pd.DataFrame, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    wb.save(path)
    print(f"  Created: {path.name}  ({len(df)} rows)")


def create_outputs(sheets: dict[str, pd.DataFrame], voice_lookup: dict[str, str]) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    print(f"=== Checkpoint 2: Building output files in '{OUTPUT_DIR}' ===\n")

    for sheet_name, df in sheets.items():
        slug = slugify(sheet_name)
        valid_df, bad_df = preprocess(df)

        save_excel(build_campaign(valid_df),               OUTPUT_DIR / f"campaign_{slug}.xlsx")
        save_excel(build_customers(valid_df, voice_lookup), OUTPUT_DIR / f"customers_{slug}.xlsx")

        if not bad_df.empty:
            save_excel(build_campaign_review(bad_df),  OUTPUT_DIR / f"campaign_{slug}.review.xlsx")
            save_excel(build_customers_review(bad_df), OUTPUT_DIR / f"customers_{slug}.review.xlsx")

    print(f"\nDone. Upload the contents of output/ to SharePoint manually.")


# ---------------------------------------------------------------------------

def main() -> None:
    if already_processed(INPUT_FILE):
        print("Input.xlsx unchanged since last run — nothing to do.")
        return

    print("=== Checkpoint 0: Reading template structures ===")
    read_template_headers(CAMPAIGN_TEMPLATE)
    read_template_headers(CUSTOMERS_TEMPLATE)

    voice_lookup = load_voice_model_lookup(CUSTOMERS_TEMPLATE)
    sheets = validate_fidelity(INPUT_FILE)
    create_outputs(sheets, voice_lookup)
    mark_processed(INPUT_FILE)
    print("Marker updated: processed/Input.xlsx.hash")


if __name__ == "__main__":
    main()
